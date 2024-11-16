#!/usr/bin/env python3
from tkinter import W
import numpy as np
import os
import pandas as pd
import xlsxwriter
from xlsxwriter import Workbook
import openpyxl
# importing PIL Module
from PIL import Image as im
import cv2
import json
import csv

def get_variables():
    # print("get_variables")
    path = os.environ["HOME"]+"/catkin_ws/src/vision_based_navigation_ttt_ml/"
    file = open(path + "variables.json")
    data = json.load(file)
    file.close()
    count = data["count"]
    count_2 = data["count_2"]
    return count, count_2
    
def update_variables(count, count_2):
    # print("update_variables")
    path = os.environ["HOME"] + "/catkin_ws/src/vision_based_navigation_ttt_ml/"
    file = open(path + "variables.json", "w")
    updated_data = {"count": count, "count_2": count_2}
    json.dump(updated_data, file)
    file.close()

def save_image(count : int, shared_path, curr_image): 
        img_name= str(count) + '.png'
        path = shared_path + img_name
        curr_image = im.fromarray(curr_image)
        print('*************************saved')
        curr_image.save(path)

def flip_images(dict):
    count, count_2 = get_variables()

    path = os.environ["HOME"]+"/catkin_ws/src/vision_based_navigation_ttt_ml/"
    path_tau_1 = path + "shape_label_tau/tau_value" 
    path_tau_2 = path + "tau_values/tau_value"    # modify
    path_folder = path + "temp/" #modify
    # load the dictionary
    csv_file_name = 'labels_file.csv'
    dict_labels = read_from_csv(path + csv_file_name)
    # print('1', dict_labels)

    folders = [file for file in sorted(os.listdir(path_folder)) if os.path.isdir(os.path.join(path_folder, file))]
    # print('folders',folders)
    for folder in folders:
        count_2 += 1
        path_images = path_folder + folder + '/'
        
        vel = folder.split('_')[4]
        path_new_folder = path_folder + 'training_images_' + str(count_2) + '_v_' + vel + '/'
        # create_folder
        os.mkdir(path_new_folder)
        images_in_folder = [f for f in sorted(os.listdir(path_images)) if f.endswith(".png")]

        # print(images_in_folder,'img_fol')
        for idx in range(len(images_in_folder)) : 
            # print('ll', images_in_folder[idx])
            try:
                # load the original input image
                image = cv2.imread(path_images + images_in_folder[idx])

                # flip the original image horizontally
                horz_img = cv2.flip(image, 1)
                # print("path",path)
                

    ##### Update tau values
                # retrive the direction from the filename
                try:    
                    ps = openpyxl.load_workbook(path_tau_1 + str(images_in_folder[idx].split('.')[0]) + '.xlsx')
                    sheet = ps['Sheet1']
                    tau_val = [sheet['A1'].value,sheet['B1'].value,sheet['C1'].value,sheet['D1'].value,sheet['E1'].value]
                    # print(count ,tau_val)
                    wb_tau =  xlsxwriter.Workbook(path_tau_1 + str(count) + ".xlsx")
                except:
                    ps = openpyxl.load_workbook(path_tau_2 + str(images_in_folder[idx].split('.')[0]) + '.xlsx')
                    sheet = ps['Sheet1']
                    tau_val = [sheet['A1'].value,sheet['B1'].value,sheet['C1'].value,sheet['D1'].value,sheet['E1'].value]
                    # print(count ,tau_val)
                    wb_tau =  xlsxwriter.Workbook(path_tau_2 + str(count) + ".xlsx")
                    
                wksheet_tau = wb_tau.add_worksheet()
                wksheet_tau.write('A1',tau_val[4])
                wksheet_tau.write('B1',tau_val[3])
                wksheet_tau.write('C1',tau_val[2])
                wksheet_tau.write('D1',tau_val[1])
                wksheet_tau.write('E1',tau_val[0])

                wb_tau.close()

    ##### update dictionary 

                image_num = str(images_in_folder[idx].split('.')[0])
                print('add dict')
                print(dict_labels[str(images_in_folder[idx].split('.')[0])])
                label = eval(dict_labels[image_num])
                count_ = str(count)
                if label == (0,0,0) or label == (1,1,1) or label == (1,1,0):
                    print(1)
                    dict_labels[count_] = str(label) # same label
                if label == (1,0,0): # left
                    print(2)
                    dict_labels[count_] = '(0,1,0)' # label right
                if label == (0,1,0): # right
                    print(3)
                    dict_labels[count_] = '(1,0,0)' # label left
                if label == (1,0,1): # left sraight
                    print(4)
                    dict_labels[count_] = '(0,1,1)' # label right straight
                if label == (0, 1, 1): # right straight
                    print(5)
                    dict_labels[count_] = '(1,0,1)' # label left straight
            
                save_dict_to_csv(dict_labels, csv_file_name)
                save_image(count, path_new_folder, horz_img)
                count += 1
                print('update var')
                update_variables(count, count_2)

            except Exception as inst:
                print('except')
                print(idx)
                print(inst)
        
def save_dict_to_csv(dictionary, dict_name):
    name = dict_name
    path = os.environ["HOME"]+"/catkin_ws/src/vision_based_navigation_ttt_ml/"
    print('save_Dicti')
    # Open the file in writing mode (no blank lines)
    with open(path + name, 'w', newline='') as f:
        # Create a CSV writer object
        writer = csv.writer(f)
        # Write one key-value tuple per row
        for row in dictionary.items():
            writer.writerow(row)

def read_from_csv(csv_filename):
    dict_from_csv = {}

    with open(csv_filename, mode='r') as inp:
        reader = csv.reader(inp)
        dict_from_csv = {rows[0]:rows[1] for rows in reader}
    return dict_from_csv

def change_brightness():
    count_new = get_variables()
    path_tau = os.environ["HOME"]+"/catkin_ws/src/vision_based_navigation_ttt_ml/tau_values/tau_value"   
    path_folder = os.environ["HOME"]+"/catkin_ws/src/vision_based_navigation_ttt_ml/training_images/"

    folders = [file for file in os.listdir(path_folder) if os.path.isdir(os.path.join(path_folder, file))]
    print('folders',folders)
    for folder in folders:
        print('folder',folder)
        # create_folder
        os.mkdir(path_folder + folder + "_brightness" )   
        path_images = os.environ["HOME"]+"/catkin_ws/src/vision_based_navigation_ttt_ml/training_images/" + folder + '/'
        images_in_folder = [f for f in listdir(path_images) if f.endswith(".png")]

        for idx in range(len(images_in_folder)) : 
            print('ll', images_in_folder[idx])
            try:
                # load the original input image
                image = cv2.imread(path_images + images_in_folder[idx],0)

                # # Displaying the image
                # cv2.imshow('image', image)
                # # waits for user to press any key
                # cv2.waitKey(0)

                # flip the original image horizontally
                alphas = [0.5, 2]
                for alpha in alphas:

                    beta = 0
                    brightness_img = cv2.convertScaleAbs(image, alpha=alpha, beta=beta)
                    path = path_folder + folder + "_brightness/" + str(count_new) + ".png"
                    print("path",path)
                    cv2.imwrite(path, brightness_img)

                    # # Displaying the image
                    # cv2.imshow("brightness", brightness_img)
                    # # waits for user to press any key
                    # cv2.waitKey(0)

                    # retrive the direction from the filename
                    ps = openpyxl.load_workbook(path_tau + str(images_in_folder[idx+1].split('.')[0]) + '.xlsx')
                    sheet = ps['Sheet1']
                    tau_val = [sheet['A1'].value,sheet['B1'].value,sheet['C1'].value,sheet['D1'].value,sheet['E1'].value]
                    print(count_new ,tau_val)
                    wb_tau =  xlsxwriter.Workbook(path_tau + str(count_new) + ".xlsx")
                    wksheet_tau = wb_tau.add_worksheet()

                #    tau_val = [tau_el, tau_l, tau_c, tau_r, tau_er]
                    wksheet_tau.write('A1',tau_val[0])
                    wksheet_tau.write('B1',tau_val[1])
                    wksheet_tau.write('C1',tau_val[2])
                    wksheet_tau.write('D1',tau_val[3])
                    wksheet_tau.write('E1',tau_val[4])

                    wb_tau.close()
                    count_new += 1

            except Exception as inst:
                print(idx)
                print(inst)
    update_variables(count_new) 

def tau():
    path = os.environ["HOME"]+"/catkin_ws/src/vision_based_navigation_ttt_ml/"
    path_tau = path + "tau_tr/" 

    files = [file for file in sorted(os.listdir(path_tau))]
    print(files)
    for file in files:
        # print(file.split('_')[1])
    # retrive the direction from the filename
        s=file.split('_')[1]
        # print(s)
        s=s[1:len(s)]
        # print(s)
        os.rename(os.path.join(path_tau,file), os.path.join(path_tau,'tau_'+ s))

def read_from_csv(csv_filename):
        dict_from_csv = {}
        with open(csv_filename, mode='r') as inp:
            reader = csv.reader(inp)
            dict_from_csv = {rows[0]:rows[1] for rows in reader}
        return dict_from_csv

def dict():
    path = os.environ["HOME"]+"/catkin_ws/src/vision_based_navigation_ttt_ml/"

    # path = os.environ["HOME"]+"/ROS-Jackal/robots/jackal/vision_based_navigation_ttt_ml/" 
    path_folder = path + "training_images/"
    csv_file_name = 'labels_file_tr.csv'
    if os.stat(path + csv_file_name).st_size == 0:
        dict_labels = {}
    else:
        dict_labels = read_from_csv(path + csv_file_name)
    print(dict_labels['26541'])
    print(eval(dict_labels['26541'])[0])
    
# change_brightness()
flip_images(1)
# tau()
# dict()
